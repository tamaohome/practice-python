# -*- coding: utf-8 -*-

"""
スプレッドシートを表示するサンプル
"""

import wx
import wx.grid
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


class MainFrame(wx.Frame):
    def __init__(self, parent, ID, title):
        wx.Frame.__init__(self, parent, title=title)
        # self.set_window_config()
        self.__create_widget()
        self.__do_layout()
        self.filepath = Path("")

    def __create_widget(self):
        self.panel = wx.Panel(self)
        self.text = wx.StaticText(self.panel, label="スプレッドシートを表示")
        self.button = wx.Button(self.panel, label="表計算ファイルを開く")
        self.button.Bind(wx.EVT_BUTTON, self.on_open_spreadsheet)

    def on_open_spreadsheet(self, event) -> None:
        with wx.FileDialog(
            self,
            "Open Spreadsheet",
            wildcard="Excel files (*.xlsx)|*.xlsx",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST,
        ) as file_dialog:
            if file_dialog.ShowModal() == wx.ID_CANCEL:
                return
            self.filepath = Path(file_dialog.GetPath())
            self.read_sheets(self.filepath)
            self.display_spreadsheet()

    def __do_layout(self):
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.text, flag=wx.ALIGN_LEFT)
        sizer.Add(self.button, flag=wx.ALIGN_CENTER | wx.TOP, border=15)
        self.panel.SetSizer(sizer)

    def read_sheets(self, path: Path | str) -> None:
        if isinstance(path, str):
            self.path = Path(path)
        self.workbook = openpyxl.load_workbook(self.path)
        self.sheets = [sheet for sheet in self.workbook.worksheets]

    def display_spreadsheet(self, position=(0, 0)) -> None:
        # self.sheet = self.workbook.worksheets[0]
        self.sheet = self.workbook.active

        if self.sheet is not type(Worksheet):
            raise ValueError(f"{self.filepath.name}: シートが無効です。")

        self.grid = wx.grid.Grid(self.panel)
        
        self.grid.CreateGrid(self.sheet.max_row, self.sheet.max_column)

        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                self.grid.SetCellValue(row - 1, col - 1, str(cell_value))

        self.grid.SetPosition(position)

        self.grid.Show()


class SpreadsheetApp(wx.App):
    def OnInit(self):
        self.frame = MainFrame(None, -1, "スプレッドシートを表示")
        self.frame.Show()
        return True


if __name__ == "__main__":
    app = SpreadsheetApp()
    app.MainLoop()
